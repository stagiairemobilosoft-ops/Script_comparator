#!/usr/bin/env python3
"""
Script de comparaison des fichiers magasins Nickel.

Sources de référence (par priorité) :
  1. Synchronisation de données automatiques site Nickel.csv
  2. nickel tache excel.xlsx

Règles métier (sources prioritaires) :
  - Fichier client = tous les établissements ouverts (feuille all_in).
  - Champs à synchroniser : nom, site web, adresse, tag, coordonnées, description,
    horaires (pos_schedule_*). Pas à partir de Date_ouverture.
  - Statut OPEN si présent dans all_in ; CLOSED si absent de all_in.
  - Magasin fermé : changement de statut en plateforme, pas de suppression.
  - Autre_numero_telephone : non synchronisé (fil CSV).

Usage:
    python compare_magasins.py compare Nickel_Gmaps_20260522.xlsx Nickel_Gmaps_20260615.xlsx
    python compare_magasins.py validate fichier.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd

STORE_ID_COLUMN = "Code_de_magasin"

FILE_COLUMNS = [
    "Code_de_magasin",
    "Tag Pays",
    "status",
    "Nom_etablissement",
    "Ligne_Adresse_1",
    "Localite",
    "pos_zipcode",
    "Latitude",
    "Longitude",
    "Siteweb",
    "pos_schedule_sunday",
    "pos_schedule_monday",
    "pos_schedule_tuesday",
    "pos_schedule_wednesday",
    "pos_schedule_thursday",
    "pos_schedule_friday",
    "pos_schedule_saturday",
    "Description_etablissement",
]

SYNC_FIELDS = [
    "Nom_etablissement",
    "Siteweb",
    "Ligne_Adresse_1",
    "Localite",
    "pos_zipcode",
    "Tag Pays",
    "Latitude",
    "Longitude",
    "Description_etablissement",
    "pos_schedule_sunday",
    "pos_schedule_monday",
    "pos_schedule_tuesday",
    "pos_schedule_wednesday",
    "pos_schedule_thursday",
    "pos_schedule_friday",
    "pos_schedule_saturday",
]

SCHEDULE_COLUMNS = [
    "pos_schedule_sunday",
    "pos_schedule_monday",
    "pos_schedule_tuesday",
    "pos_schedule_wednesday",
    "pos_schedule_thursday",
    "pos_schedule_friday",
    "pos_schedule_saturday",
]

OPTIONAL_COLUMNS = [
    "Numero_telephone",
    "Horaires_ouverture_exceptionnels",
]

DAY_SCHEDULE_CODES = [
    ("pos_schedule_monday", "MO"),
    ("pos_schedule_tuesday", "TU"),
    ("pos_schedule_wednesday", "WE"),
    ("pos_schedule_thursday", "TH"),
    ("pos_schedule_friday", "FR"),
    ("pos_schedule_saturday", "SA"),
    ("pos_schedule_sunday", "SU"),
]

RESULT_COLUMNS = [
    "SHOP_CODE",
    "ACTION",
    "STATUS",
    "LOCATION_NAME",
    "LOCATION_LANGUAGE",
    "PHONE_NUMBER",
    "EMAIL",
    "WEBSITE_URL",
    "ADR_STREET_NAME",
    "ADR_STREET_NUMBER",
    "ADR_ZIPCODE",
    "ADR_CITY",
    "ADR_COUNTRY",
    "ADR_LATITUDE",
    "ADR_LONGITUDE",
    "REGULAR_HOURS",
    "SPECIAL_HOURS (OPTIONAL)",
]

CLIENT_COLUMN_RENAME = {
    "Pays": "Tag Pays",
    "Ligne_adresse1": "Ligne_Adresse_1",
}

OUTPUT_DIR = Path("output")
OUTPUT_RESULTAT = OUTPUT_DIR / "resultat"
OUTPUT_LOG = OUTPUT_DIR / "log"
DEFAULT_COMPARE_OUTPUT = OUTPUT_RESULTAT / "fichier_synchronisation.xlsx"
DEFAULT_VALIDATE_OUTPUT = OUTPUT_LOG / "problemes_qualite.xlsx"
STATUS_OPEN = "OPEN"
STATUS_CLOSED = "CLOSED"

# Actions du flow IT (comparaison mois N vs mois N-1)
ACTION_NO_CHANGE = "no_change"
ACTION_UPDATE = "update"
ACTION_CREATE = "create"
ACTION_CLOSE = "close_permanently"
ACTION_COLUMN = "action"

EXPORT_ACTION_MAP = {
    ACTION_UPDATE: "UPDATE",
    ACTION_CREATE: "CREATE",
    ACTION_CLOSE: "UPDATE",
}


class LoadError(Exception):
    pass


class ProgressReporter:
    """Affiche la progression sur stderr (ligne courante / total)."""

    def __init__(
        self,
        label: str,
        total: int,
        *,
        enabled: bool = True,
        interval: int = 500,
    ) -> None:
        self.label = label
        self.total = max(total, 1)
        self.enabled = enabled
        self.interval = max(interval, 1)
        self._last_shown = 0

    def update(self, current: int) -> None:
        if not self.enabled or current < 1:
            return
        if current != self.total and current % self.interval != 0 and current != 1:
            return
        self._last_shown = current
        percent = int(current * 100 / self.total)
        line = f"      {self.label} : {current} / {self.total} ({percent}%)"
        sys.stderr.write(f"\r{line:<70}")
        sys.stderr.flush()

    def finish(self, detail: str = "") -> None:
        if not self.enabled:
            return
        if self._last_shown != self.total:
            sys.stderr.write(f"\r      {self.label} : {self.total} / {self.total}   ")
        sys.stderr.write("\n")
        if detail:
            sys.stderr.write(f"      {detail}\n")
        sys.stderr.flush()


def _print_stage(step: int, total: int, message: str, *, enabled: bool = True) -> None:
    if enabled:
        print(f"[{step}/{total}] {message}", file=sys.stderr, flush=True)


def _print_step_done(message: str, *, enabled: bool = True) -> None:
    if enabled:
        print(f"      {message}", file=sys.stderr, flush=True)


def normalize_cell(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return ""
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        return text[:-2]
    return text


def _is_client_format(columns: list[str]) -> bool:
    if STORE_ID_COLUMN not in columns:
        return False
    if "Ligne_adresse1" in columns or "Ligne_Adresse_1" in columns:
        return True
    return "Nom_etablissement" in columns and ("Pays" in columns or "Tag Pays" in columns)


def _find_data_sheet(path: Path, sheet_names: list[str]) -> str:
    for name in sheet_names:
        if "all_in" in name.lower():
            return name

    for name in sheet_names:
        header = pd.read_excel(path, sheet_name=name, nrows=0)
        columns = [str(col).strip() for col in header.columns]
        if _is_client_format(columns):
            row_count = len(pd.read_excel(path, sheet_name=name, usecols=[0]))
            if row_count > 0:
                return name

    raise LoadError(
        "Aucune feuille de magasins trouvée (all_in ou format Nickel Gmaps). "
        f"Feuilles disponibles : {', '.join(sheet_names)}"
    )


def _normalize_raw_dataframe(df: pd.DataFrame, statut: str = "") -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    if "Statut" in df.columns and "status" not in df.columns:
        df = df.rename(columns={"Statut": "status"})

    if _is_client_format(list(df.columns)):
        df = df.rename(columns=CLIENT_COLUMN_RENAME)
        if not statut:
            statut = STATUS_OPEN

    if "status" not in df.columns:
        df["status"] = statut

    missing = [col for col in FILE_COLUMNS if col not in df.columns]
    if missing:
        raise LoadError(f"Colonnes manquantes : {', '.join(missing)}")

    for col in OPTIONAL_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[[*FILE_COLUMNS, *OPTIONAL_COLUMNS]].copy()
    ids = df[STORE_ID_COLUMN].astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
    df[STORE_ID_COLUMN] = ids
    return df


def load_store_file(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise LoadError(f"Fichier introuvable : {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        workbook = pd.ExcelFile(path)
        first_sheet = pd.read_excel(path, sheet_name=workbook.sheet_names[0], nrows=0)
        columns = [str(col).strip() for col in first_sheet.columns]

        if _is_client_format(columns) or any(
            _is_client_format([str(c).strip() for c in pd.read_excel(path, sheet_name=n, nrows=0).columns])
            for n in workbook.sheet_names
        ):
            sheet_name = _find_data_sheet(path, workbook.sheet_names)
            df = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
            return _normalize_raw_dataframe(df, statut=STATUS_OPEN)

        if set(FILE_COLUMNS).issubset(columns):
            df = pd.read_excel(path, dtype=str)
            return _normalize_raw_dataframe(df)

        raise LoadError(
            "Format de fichier non reconnu. Attendu : export Nickel Gmaps (feuille all_in)."
        )

    if suffix == ".csv":
        df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
        return _normalize_raw_dataframe(df)

    raise LoadError(f"Format non supporté : {suffix}. Utilisez .xlsx, .xls ou .csv")


@dataclass
class ComparisonResult:
    modified: pd.DataFrame = field(default_factory=pd.DataFrame)
    added: pd.DataFrame = field(default_factory=pd.DataFrame)
    removed: pd.DataFrame = field(default_factory=pd.DataFrame)
    unchanged: pd.DataFrame = field(default_factory=pd.DataFrame)
    changed_fields: pd.DataFrame = field(default_factory=pd.DataFrame)
    reference_total: int = 0
    incoming_total: int = 0

    @property
    def unchanged_count(self) -> int:
        return len(self.unchanged)


def compare_store_files(
    reference: pd.DataFrame,
    incoming: pd.DataFrame,
    *,
    show_progress: bool = True,
) -> ComparisonResult:
    ref = _index_by_store(reference)
    inc = _index_by_store(incoming)

    ref_ids = set(ref.index)
    inc_ids = set(inc.index)
    added_ids = inc_ids - ref_ids
    removed_ids = ref_ids - inc_ids
    common_ids = ref_ids & inc_ids

    changed_rows: list[pd.Series] = []
    unchanged_rows: list[pd.Series] = []
    field_changes: list[dict[str, str]] = []

    common_list = sorted(common_ids)
    progress = ProgressReporter("Comparaison", len(common_list), enabled=show_progress)
    for index, store_id in enumerate(common_list, start=1):
        ref_row = ref.loc[store_id]
        inc_row = inc.loc[store_id]
        changed = [
            name
            for name in SYNC_FIELDS
            if normalize_cell(ref_row[name]) != normalize_cell(inc_row[name])
        ]
        if changed:
            changed_rows.append(inc.loc[store_id])
            for field_name in changed:
                field_changes.append(
                    {
                        STORE_ID_COLUMN: store_id,
                        "champ": field_name,
                        "ancienne_valeur": normalize_cell(ref_row[field_name]),
                        "nouvelle_valeur": normalize_cell(inc_row[field_name]),
                    }
                )
        else:
            unchanged_rows.append(inc.loc[store_id])
        progress.update(index)
    progress.finish(
        f"{len(changed_rows)} modification(s), {len(unchanged_rows)} inchange(s)"
    )

    added = incoming[incoming[STORE_ID_COLUMN].isin(added_ids)].copy()
    added["status"] = STATUS_OPEN

    removed = reference[reference[STORE_ID_COLUMN].isin(removed_ids)].copy()
    removed["status"] = STATUS_CLOSED

    return ComparisonResult(
        modified=_rows_to_dataframe(changed_rows, incoming.columns),
        added=added,
        removed=removed,
        unchanged=_rows_to_dataframe(unchanged_rows, incoming.columns),
        changed_fields=pd.DataFrame(field_changes),
        reference_total=len(ref_ids),
        incoming_total=len(inc_ids),
    )


def _issue(store_id: str, type_probleme: str, detail: str, gravite: str) -> dict[str, str]:
    return {
        STORE_ID_COLUMN: store_id,
        "type_probleme": type_probleme,
        "detail": detail,
        "gravite": gravite,
    }


def find_data_quality_issues(df: pd.DataFrame, *, show_progress: bool = True) -> pd.DataFrame:
    """Rapport CS : erreurs bloquantes et avertissements informatifs."""
    issues: list[dict[str, str]] = []

    duplicated = df[df.duplicated(subset=[STORE_ID_COLUMN], keep=False)]
    for store_id in duplicated[STORE_ID_COLUMN].unique():
        issues.append(
            _issue(
                store_id,
                "doublon",
                "Code magasin présent plusieurs fois dans le fichier",
                "erreur",
            )
        )

    progress = ProgressReporter("Controle qualite", len(df), enabled=show_progress)
    for index, (_, row) in enumerate(df.iterrows(), start=1):
        store_id = normalize_cell(row[STORE_ID_COLUMN])
        for field, label in {
            "Nom_etablissement": "nom manquant",
            "Ligne_Adresse_1": "adresse manquante",
            "Localite": "ville manquante",
            "pos_zipcode": "code postal manquant",
            "Latitude": "latitude manquante",
            "Longitude": "longitude manquante",
        }.items():
            if not normalize_cell(row[field]):
                issues.append(_issue(store_id, "champ_manquant", label, "erreur"))

        lat = _to_float(row["Latitude"])
        lon = _to_float(row["Longitude"])
        if lat is not None and not -90 <= lat <= 90:
            issues.append(
                _issue(store_id, "coordonnee_invalide", f"latitude hors plage : {lat}", "erreur")
            )
        if lon is not None and not -180 <= lon <= 180:
            issues.append(
                _issue(store_id, "coordonnee_invalide", f"longitude hors plage : {lon}", "erreur")
            )

        # Horaire vide : avertissement CS seulement (non bloquant, pas de règle explicite dans les sources prioritaires)
        if not any(normalize_cell(row[col]) for col in SCHEDULE_COLUMNS):
            issues.append(
                _issue(
                    store_id,
                    "horaire_manquant",
                    "aucun horaire renseigné (à vérifier avec le client si besoin)",
                    "avertissement",
                )
            )

        website = normalize_cell(row["Siteweb"])
        if website and not website.startswith(("http://", "https://")):
            issues.append(
                _issue(store_id, "site_web", "URL sans schéma http/https", "avertissement")
            )
        progress.update(index)
    progress.finish()

    columns = [STORE_ID_COLUMN, "type_probleme", "detail", "gravite"]
    if not issues:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(issues).drop_duplicates()


def deduplicate_stores(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Garde la dernière occurrence par code magasin. Retourne (unique, doublons)."""
    df = df.copy()
    df[STORE_ID_COLUMN] = df[STORE_ID_COLUMN].map(normalize_cell)
    duplicated_mask = df[STORE_ID_COLUMN].duplicated(keep=False)
    duplicates = df[duplicated_mask].sort_values(STORE_ID_COLUMN)
    unique = df.drop_duplicates(subset=[STORE_ID_COLUMN], keep="last")
    return unique, duplicates


def _index_by_store(df: pd.DataFrame) -> pd.DataFrame:
    indexed = df.copy()
    indexed[STORE_ID_COLUMN] = indexed[STORE_ID_COLUMN].map(normalize_cell)
    return indexed.set_index(STORE_ID_COLUMN, drop=False)


def _rows_to_dataframe(rows: list[pd.Series], columns: pd.Index) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows).reset_index(drop=True)


def _to_float(value: object) -> float | None:
    text = normalize_cell(value)
    if not text:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


def _cell_display_length(value: object) -> int:
    if value is None:
        return 0
    text = str(value)
    if "\n" in text:
        return max((len(line) for line in text.split("\n")), default=0)
    return len(text)


def _autofit_columns(path: Path, max_width: int = 60) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            width = max((_cell_display_length(cell.value) for cell in column_cells), default=0)
            sheet.column_dimensions[letter].width = min(width + 2, max_width)
    workbook.save(path)


def _write_excel(path: Path, df: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, engine="openpyxl")
    _autofit_columns(path)


def _write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
    _autofit_columns(path)


def _ensure_output_dirs() -> None:
    OUTPUT_RESULTAT.mkdir(parents=True, exist_ok=True)
    OUTPUT_LOG.mkdir(parents=True, exist_ok=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _timestamped_log_path(prefix: str, stamp: str | None = None) -> Path:
    return OUTPUT_LOG / f"{prefix}_{stamp or _timestamp()}.log"


def _timestamped_analysis_path(stamp: str | None = None) -> Path:
    return OUTPUT_LOG / f"analyse_{stamp or _timestamp()}.xlsx"


def _write_log(path: Path, lines: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _sync_columns() -> list[str]:
    return [ACTION_COLUMN, *FILE_COLUMNS, *OPTIONAL_COLUMNS]


def _with_action(df: pd.DataFrame, action: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=_sync_columns())
    out = df.copy()
    out.insert(0, ACTION_COLUMN, action)
    return out


def split_address(address: str) -> tuple[str, str]:
    """Sépare nom de rue et numéro (fin ou début de ligne)."""
    address = address.strip()
    if not address:
        return "", ""

    match = re.match(r"^(.*?)[\s,]+(\d+\w?)$", address)
    if match:
        return match.group(1).strip(), match.group(2).strip()

    match = re.match(r"^(\d+\w?)\s+(.+)$", address)
    if match:
        return match.group(2).strip(), match.group(1).strip()

    return address, ""


def _join_regular_hours(row: pd.Series) -> str:
    """Format CLIENT_EXPORT : MO:"10:00"-"12:30";"13:00"-"19:30","""
    day_parts: list[str] = []
    for column, day_code in DAY_SCHEDULE_CODES:
        value = normalize_cell(row.get(column, ""))
        if not value:
            continue
        slots: list[str] = []
        for slot in value.split(","):
            slot = slot.strip()
            if "-" not in slot:
                continue
            start, end = slot.split("-", 1)
            slots.append(f'"{start.strip()}"-"{end.strip()}"')
        if slots:
            day_parts.append(f'{day_code}:{";".join(slots)}')

    if not day_parts:
        return ""

    lines: list[str] = []
    for index, part in enumerate(day_parts):
        lines.append(f"{part}," if index < len(day_parts) - 1 else part)
    return "\n".join(lines)


def to_result_layout(sync_df: pd.DataFrame, *, show_progress: bool = True) -> pd.DataFrame:
    """Réorganise les colonnes selon le modèle CLIENT_EXPORT."""
    if sync_df.empty:
        return pd.DataFrame(columns=RESULT_COLUMNS)

    rows: list[dict[str, str]] = []
    progress = ProgressReporter("Export Filezilla", len(sync_df), enabled=show_progress)
    for index, (_, row) in enumerate(sync_df.iterrows(), start=1):
        action = normalize_cell(row[ACTION_COLUMN])
        country = normalize_cell(row["Tag Pays"])
        street, number = split_address(normalize_cell(row["Ligne_Adresse_1"]))
        rows.append(
            {
                "SHOP_CODE": normalize_cell(row[STORE_ID_COLUMN]),
                "ACTION": EXPORT_ACTION_MAP.get(action, action.upper()),
                "STATUS": normalize_cell(row["status"]),
                "LOCATION_NAME": normalize_cell(row["Nom_etablissement"]),
                "LOCATION_LANGUAGE": country,
                "PHONE_NUMBER": normalize_cell(row.get("Numero_telephone", "")),
                "EMAIL": "",
                "WEBSITE_URL": normalize_cell(row["Siteweb"]),
                "ADR_STREET_NAME": street,
                "ADR_STREET_NUMBER": number,
                "ADR_ZIPCODE": normalize_cell(row["pos_zipcode"]),
                "ADR_CITY": normalize_cell(row["Localite"]),
                "ADR_COUNTRY": country,
                "ADR_LATITUDE": normalize_cell(row["Latitude"]),
                "ADR_LONGITUDE": normalize_cell(row["Longitude"]),
                "REGULAR_HOURS": _join_regular_hours(row),
                "SPECIAL_HOURS (OPTIONAL)": normalize_cell(
                    row.get("Horaires_ouverture_exceptionnels", "")
                ),
            }
        )
        progress.update(index)
    progress.finish()
    return pd.DataFrame(rows, columns=RESULT_COLUMNS)


def build_sync_file(result: ComparisonResult) -> pd.DataFrame:
    """
    Fichier unique pour Filezilla / batch de synchro (flow IT).

    Contient uniquement les établissements à mettre à jour, créer ou fermer.
    Les magasins sans mise à jour sont exclus (liste séparée dans le résumé console).
    """
    parts = [
        _with_action(result.modified, ACTION_UPDATE),
        _with_action(result.added, ACTION_CREATE),
        _with_action(result.removed, ACTION_CLOSE),
    ]
    non_empty = [part for part in parts if not part.empty]
    if not non_empty:
        return pd.DataFrame(columns=_sync_columns())
    return pd.concat(non_empty, ignore_index=True)


def _build_analysis_summary(
    result: ComparisonResult,
    previous: Path,
    incoming: Path,
    resultat: Path,
    dup_count: int,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"indicateur": "date", "valeur": datetime.now().isoformat(timespec="seconds")},
            {"indicateur": "fichier_mois_precedent", "valeur": str(previous.resolve())},
            {"indicateur": "fichier_nouveau", "valeur": str(incoming.resolve())},
            {"indicateur": "fichier_resultat", "valeur": str(resultat.resolve())},
            {"indicateur": "total_reference", "valeur": result.reference_total},
            {"indicateur": "total_entrant", "valeur": result.incoming_total},
            {"indicateur": ACTION_NO_CHANGE, "valeur": result.unchanged_count},
            {"indicateur": ACTION_UPDATE, "valeur": len(result.modified)},
            {"indicateur": ACTION_CREATE, "valeur": len(result.added)},
            {"indicateur": ACTION_CLOSE, "valeur": len(result.removed)},
            {"indicateur": "lignes_filezilla", "valeur": len(result.modified) + len(result.added) + len(result.removed)},
            {"indicateur": "doublons_entrant", "valeur": dup_count},
        ]
    )


def _write_analysis_workbook(
    path: Path,
    result: ComparisonResult,
    previous: Path,
    incoming: Path,
    resultat: Path,
    dup_count: int,
    inc_duplicates: pd.DataFrame,
    quality: pd.DataFrame,
) -> None:
    _write_workbook(
        path,
        {
            "resume": _build_analysis_summary(result, previous, incoming, resultat, dup_count),
            "no_change": _with_action(result.unchanged, ACTION_NO_CHANGE),
            "update": _with_action(result.modified, ACTION_UPDATE),
            "create": _with_action(result.added, ACTION_CREATE),
            "close_permanently": _with_action(result.removed, ACTION_CLOSE),
            "champs_modifies": result.changed_fields,
            "doublons": inc_duplicates,
            "qualite": quality,
        },
    )


def cmd_compare(args: argparse.Namespace) -> int:
    show_progress = not args.quiet
    previous_path = Path(args.previous)
    incoming_path = Path(args.incoming)

    try:
        _print_stage(1, 5, f"Chargement {previous_path.name}...", enabled=show_progress)
        ref_df, _ = deduplicate_stores(load_store_file(previous_path))
        _print_step_done(f"{len(ref_df)} magasin(s) charges", enabled=show_progress)

        _print_stage(2, 5, f"Chargement {incoming_path.name}...", enabled=show_progress)
        inc_raw = load_store_file(incoming_path)
        inc_df, inc_duplicates = deduplicate_stores(inc_raw)
        _print_step_done(f"{len(inc_df)} magasin(s) charges", enabled=show_progress)

        _print_stage(3, 5, "Comparaison des magasins...", enabled=show_progress)
        result = compare_store_files(ref_df, inc_df, show_progress=show_progress)
    except LoadError as exc:
        print(f"Erreur : {exc}", file=sys.stderr)
        return 1

    _ensure_output_dirs()
    output_path = Path(args.output) if args.output else DEFAULT_COMPARE_OUTPUT
    if not output_path.is_absolute() and output_path.parent == Path("."):
        output_path = OUTPUT_RESULTAT / output_path.name

    _print_stage(4, 5, "Generation du fichier Filezilla...", enabled=show_progress)
    sync_file = build_sync_file(result)
    _write_excel(output_path, to_result_layout(sync_file, show_progress=show_progress))
    _print_step_done(f"{len(sync_file)} ligne(s) ecrites", enabled=show_progress)

    _print_stage(5, 5, "Controle qualite et rapport d'analyse...", enabled=show_progress)
    quality = find_data_quality_issues(inc_raw, show_progress=show_progress)
    dup_count = inc_duplicates[STORE_ID_COLUMN].nunique() if not inc_duplicates.empty else 0

    run_stamp = _timestamp()
    analysis_path = _timestamped_analysis_path(run_stamp)
    _write_analysis_workbook(
        analysis_path, result, previous_path, incoming_path, output_path, dup_count, inc_duplicates, quality
    )
    _print_step_done("Rapport d'analyse enregistre", enabled=show_progress)
    if show_progress:
        print(file=sys.stderr)

    log_lines = [
        f"date={datetime.now().isoformat(timespec='seconds')}",
        f"commande=compare",
        f"mois_precedent={previous_path.resolve()}",
        f"nouveau={incoming_path.resolve()}",
        f"resultat={output_path.resolve()}",
        f"analyse={analysis_path.resolve()}",
        f"total_reference={result.reference_total}",
        f"total_entrant={result.incoming_total}",
        f"{ACTION_NO_CHANGE}={result.unchanged_count}",
        f"{ACTION_UPDATE}={len(result.modified)}",
        f"{ACTION_CREATE}={len(result.added)}",
        f"{ACTION_CLOSE}={len(result.removed)}",
        f"lignes_filezilla={len(sync_file)}",
        f"doublons={dup_count}",
    ]
    log_path = _timestamped_log_path("compare", run_stamp)
    _write_log(log_path, log_lines)

    print("Classification terminée (flow IT).")
    if dup_count:
        print(f"  Doublons entrant : {dup_count} code(s) en double (dernière ligne conservée)")
    print(f"  Mois précédent             : {result.reference_total} magasins")
    print(f"  Nouveau fichier            : {result.incoming_total} magasins")
    print(f"  {ACTION_NO_CHANGE}       : {result.unchanged_count} (exclus du fichier)")
    print(f"  {ACTION_UPDATE}          : {len(result.modified)} (status {STATUS_OPEN})")
    print(f"  {ACTION_CREATE}          : {len(result.added)} (status {STATUS_OPEN})")
    print(f"  {ACTION_CLOSE}           : {len(result.removed)} (status {STATUS_CLOSED})")
    print(f"  Lignes dans le fichier Filezilla : {len(sync_file)}")
    print(f"Résultat : {output_path.resolve()}")
    print(f"Analyse  : {analysis_path.resolve()}")
    print(f"Log      : {log_path.resolve()}")
    return 0


def cmd_validate(args: argparse.Namespace) -> int:
    try:
        df = load_store_file(Path(args.fichier))
    except LoadError as exc:
        print(f"Erreur : {exc}", file=sys.stderr)
        return 1

    _ensure_output_dirs()
    issues = find_data_quality_issues(df)
    output = Path(args.output)
    _write_excel(output, issues)

    log_path = _timestamped_log_path("validate")
    _write_log(
        log_path,
        [
            f"date={datetime.now().isoformat(timespec='seconds')}",
            "commande=validate",
            f"fichier={Path(args.fichier).resolve()}",
            f"problemes={len(issues)}",
            f"rapport={output.resolve()}",
        ],
    )
    print(f"{len(issues)} problème(s) détecté(s). Rapport : {output.resolve()}")
    print(f"Log : {log_path.resolve()}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Comparaison de fichiers magasins Nickel pour la synchronisation FTP."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    compare_parser = subparsers.add_parser(
        "compare",
        help="Compare le fichier du mois précédent avec le nouveau fichier Nickel",
    )
    compare_parser.add_argument(
        "previous",
        help="Fichier Nickel du mois précédent (déjà synchronisé)",
    )
    compare_parser.add_argument("incoming", help="Nouveau fichier Nickel reçu")
    compare_parser.add_argument(
        "-o",
        "--output",
        default=str(DEFAULT_COMPARE_OUTPUT),
        help="Fichier Excel pour Filezilla (défaut: output/resultat/fichier_synchronisation.xlsx)",
    )
    compare_parser.add_argument(
        "-q",
        "--quiet",
        action="store_true",
        help="Masquer l'indicateur de progression",
    )
    compare_parser.set_defaults(func=cmd_compare)

    validate_parser = subparsers.add_parser(
        "validate", help="Contrôle la qualité des données d'un fichier (usage CS)"
    )
    validate_parser.add_argument("fichier", help="Fichier Nickel à contrôler")
    validate_parser.add_argument(
        "-o",
        "--output",
        default=str(DEFAULT_VALIDATE_OUTPUT),
        help="Rapport Excel dans output/log (défaut: output/log/problemes_qualite.xlsx)",
    )
    validate_parser.set_defaults(func=cmd_validate)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
