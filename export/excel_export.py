from __future__ import annotations
import argparse
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import pandas as pd

from constants.constant import (
    STORE_ID_COLUMN,
    RESULT_COLUMNS,
    ACTION_UPDATE,
    ACTION_CLOSE,
    ACTION_CREATE,
    DAY_SCHEDULE_CODES,
    ACTION_COLUMN,
    EXPORT_ACTION_MAP,
    FILE_COLUMNS,
    OPTIONAL_COLUMNS,
)
from comparison.progress import *
from comparison.utils import normalize_cell
from comparison.compare import ComparisonResult

def _cell_display_length(value: object) -> int:
    if value is None:
        return 0
    text = str(value)
    if "\n" in text:
        return max((len(line) for line in text.split("\n")), default=0)
    return len(text)

def _sync_columns() -> list[str]:
    return [ACTION_COLUMN, *FILE_COLUMNS, *OPTIONAL_COLUMNS]

def with_action(df: pd.DataFrame, action: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=_sync_columns())
    out = df.copy()
    out.insert(0, ACTION_COLUMN, action)
    return out

def _autofit_columns(path: Path, max_width: int = 60) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            width = max((_cell_display_length(cell.value) for cell in column_cells), default=0)
            sheet.column_dimensions[letter].width = min(width + 2, max_width)
    workbook.save(path)

def write_excel(path: Path, df: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False, engine="openpyxl")
    _autofit_columns(path)

def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
    _autofit_columns(path)

def split_address(address: str) -> tuple[str, str]:
    """Sépare nom de rue et numéro (fin ou début de ligne)."""
    address = address.strip()
    if not address:
        return "", ""

    match = re.match(r"^(.*?)[\s,]+(\d+\w?)$", address)
    if match:
        return match.group(1).strip(), match.group(2).strip()

    match = re.match(r"^(\d+\w?)\s+(.+)$", address)
    if match:
        return match.group(2).strip(), match.group(1).strip()

    return address, ""

def _join_regular_hours(row: pd.Series) -> str:
    """Format CLIENT_EXPORT : MO:"10:00"-"12:30";"13:00"-"19:30","""
    day_parts: list[str] = []
    for column, day_code in DAY_SCHEDULE_CODES:
        value = normalize_cell(row.get(column, ""))
        if not value:
            continue
        slots: list[str] = []
        for slot in value.split(","):
            slot = slot.strip()
            if "-" not in slot:
                continue
            start, end = slot.split("-", 1)
            slots.append(f'"{start.strip()}"-"{end.strip()}"')
        if slots:
            day_parts.append(f'{day_code}:{";".join(slots)}')

    if not day_parts:
        return ""

    lines: list[str] = []
    for index, part in enumerate(day_parts):
        lines.append(f"{part}," if index < len(day_parts) - 1 else part)
    return "\n".join(lines)



def to_result_layout(sync_df: pd.DataFrame, *, show_progress: bool = True) -> pd.DataFrame:
    """Réorganise les colonnes selon le modèle CLIENT_EXPORT."""
    if sync_df.empty:
        return pd.DataFrame(columns=RESULT_COLUMNS)

    rows: list[dict[str, str]] = []
    progress = ProgressReporter("Export Filezilla", len(sync_df), enabled=show_progress)
    for index, (_, row) in enumerate(sync_df.iterrows(), start=1):
        action = normalize_cell(row[ACTION_COLUMN])
        country = normalize_cell(row["Tag Pays"])
        street, number = split_address(normalize_cell(row["Ligne_Adresse_1"]))
        rows.append(
            {
                "SHOP_CODE": normalize_cell(row[STORE_ID_COLUMN]),
                "ACTION": EXPORT_ACTION_MAP.get(action, action.upper()),
                "STATUS": normalize_cell(row["status"]),
                "LOCATION_NAME": normalize_cell(row["Nom_etablissement"]),
                "LOCATION_LANGUAGE": country,
                "PHONE_NUMBER": normalize_cell(row.get("Numero_telephone", "")),
                "EMAIL": "",
                "WEBSITE_URL": normalize_cell(row["Siteweb"]),
                "ADR_STREET_NAME": street,
                "ADR_STREET_NUMBER": number,
                "ADR_ZIPCODE": normalize_cell(row["pos_zipcode"]),
                "ADR_CITY": normalize_cell(row["Localite"]),
                "ADR_COUNTRY": country,
                "ADR_LATITUDE": normalize_cell(row["Latitude"]),
                "ADR_LONGITUDE": normalize_cell(row["Longitude"]),
                "REGULAR_HOURS": _join_regular_hours(row),
                "SPECIAL_HOURS (OPTIONAL)": normalize_cell(
                    row.get("Horaires_ouverture_exceptionnels", "")
                ),
            }
        )
        progress.update(index)
    progress.finish()
    return pd.DataFrame(rows, columns=RESULT_COLUMNS)



def build_sync_file(result: ComparisonResult) -> pd.DataFrame:
    """
    Fichier unique pour Filezilla / batch de synchro (flow IT).

    Contient uniquement les établissements à mettre à jour, créer ou fermer.
    Les magasins sans mise à jour sont exclus (liste séparée dans le résumé console).
    """
    parts = [
        with_action(result.modified, ACTION_UPDATE),
        with_action(result.added, ACTION_CREATE),
        with_action(result.removed, ACTION_CLOSE),
    ]
    non_empty = [part for part in parts if not part.empty]
    if not non_empty:
        return pd.DataFrame(columns=_sync_columns())
    return pd.concat(non_empty, ignore_index=True)
